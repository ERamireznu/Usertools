#v00: 06/07/25
#v01: 10/07/25
#v02: 07/11/25 tkinter gui
#-------------------------------
import tkinter as tk
from tkinter import Toplevel
import datetime
ahora = str(datetime.datetime.now())

def path_parts(pat0):   #reads a path, separates dir and file
    if pat0[-1]=='\\': #a dir ending in \
        return (pat0[:-1], '')
    Lfname = []
    for i in range(-1, -len(pat0)-1,-1):
        if pat0[i]=='\\':
            break
        else:
            Lfname.insert(0, pat0[i])
    fname = ''.join(Lfname)
    fubic = pat0[:len(pat0)+i]
    if '.' not in fname:  #it's a dir
        fname = ''
        fubic = pat0
    return (fubic, fname)

import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
tt, TT, ff, FF = True, True, False, False
global path_def
path_def = "C:\\Users\\esteb\\xls_exp.xlsx"

#--------------------------------
def window00(text00, text01, fx00):
    global ent01, var01
    vent_fd = tk.Toplevel() #Tk()
    vent_fd.geometry("400x130")
    vent_fd.title(text00)
    lab01 = tk.Label(vent_fd, text=f"Insert file path (n: default --> {path_def})", anchor="w")
    ent01 = tk.Entry(vent_fd, width=60)

    var01 = tk.BooleanVar(value=False)  # starts unchecked
    chk01 = tk.Checkbutton(vent_fd, variable=var01)
    lab02 = tk.Label(vent_fd, text= text01, anchor="w")
    but_fd0 = tk.Button(vent_fd, text="Submit", command = fx00)

    lab01.place(x=10, y=10)    
    ent01.place(x=10, y=35)  
    chk01.place(x=10, y=60)
    lab02.place(x=30, y=60)
    but_fd0.place(x=10, y=90)

    vent_fd.mainloop()

def filedata_show():
    if ent01.get().lower() == 'n':
        path = path_def                    
    else:
        path = ent01.get()   

    path_exp = path_parts(path)
    fubic_, fname_ = path_exp[0], path_exp[1]
    wb = load_workbook(path, data_only=False)  #considers cells with formulas
    Res = []
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        used_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
        Res.append((hoja, used_cells))
    Tot_usedcells = sum([x[1] for x in Res])

    print(f"File  : {fname_}")
    print(f"Ubic  : {fubic_}")
    print(f"#Shts : {len(wb.sheetnames)}")
    print(f"#cells: {Tot_usedcells} (total)")

    if var01.get():
        print('Sheets ----------')
        Res.insert(0,('Name', '#cells used:'))
        Res_lens = [(len(x[0]), len(str(x[1]))) for x in Res]
        max0, max1 = max([x[0] for x in Res_lens]), max([x[1] for x in Res_lens])
        Res2 = [(x[0]+(max0-len(x[0])+1)*' ', str(x[1])+(max1-len(str(x[1]))+1)*' ') for x in Res]
        for x in Res2:
            print(f"{x[0]} {x[1]}")
        
    print('-'*20)

def window01(text00, text01, fx00):
    global ent10, ent11
    vent_fd = tk.Toplevel() #Tk()
    vent_fd.geometry("400x160")
    vent_fd.title(text00)
    lab01 = tk.Label(vent_fd, text=f"Insert dir path (n: default --> {path_parts(path_def)[0]})", anchor="w")
    ent10 = tk.Entry(vent_fd, width=60)

    lab02 = tk.Label(vent_fd, text= text01, anchor="w")
    ent11 = tk.Entry(vent_fd, width=60)
    but_fd0 = tk.Button(vent_fd, text="Submit", command = fx00)

    lab01.place(x=10, y=10)    
    ent10.place(x=10, y=35)  
    lab02.place(x=10, y=70)
    ent11.place(x=10, y=95)
    but_fd0.place(x=10, y=125)

    vent_fd.mainloop()

global sh_list
def user_entry():
    if ent10.get().lower() == 'n':
        path = path_def                    
    else:
        path = ent10.get()   
    path_exp = path_parts(path)
    fubic_ = path_exp[0]

    en_user = ent11.get()
    sh_list = en_user.split(',')  #separated with commas?
    if len(sh_list)==1:
        sh_list = en_user.split()  #no sep w/commas: now split with space
    elif len(sh_list)>1:
        sh_list = [(x.lstrip()).rstrip() for x in sh_list]
    return sh_list, fubic_
    
def create_files_show():
    user_ = user_entry()
    sh_list00, fubic_00 = user_[0], user_[1]
    # make files from a list:
    for nom in sh_list00:    
        wb = openpyxl.Workbook()
        fpath = fubic_00 + '\\' + str(nom) +".xlsx"
        wb.save(fpath)               

def create_sheets_show():
    user_ = user_entry()
    sh_list00, fubic_00 = user_[0], user_[1]
    # make sheets from a list:
    wb = openpyxl.Workbook()
    sheet = wb.active        
    for i in range(len(sh_list00)):
        wb.create_sheet(index = i, title = sh_list00[i])
    import datetime
    now = datetime.datetime.now()
    dt = now.strftime('%d%m%y_%H%M')
    nom = 'new_' + dt +'.xlsx'        
    wb.save(fubic_00 + '\\' + nom)
    print(f"file '{nom}' created at {fubic_00}")

def filedata_table(event):
    window00("File Data", "Show sheets details", filedata_show)

def create_files(event):
    window01("Create files", "Enter files names (Ex1: a,b,c || Ex2: aa bb cc)", create_files_show)

def create_sheets(event):
    window01("Create sheets", "Enter sheets names, to a new file (Ex1: a,b,c || Ex2: aa bb cc)", create_sheets_show)    

    
#start--------------------------------------------------------
print(ahora[:19])
root = tk.Tk()
root.geometry("140x120")
root.title("Excel tools")

lab00 = tk.Label(root, text="Excel tools", font=("Arial",9,"bold"), anchor="w")
but01 = tk.Button(root, text="File Data")
but02 = tk.Button(root, text="Create Files")
but03 = tk.Button(root, text="Create Sheets")
lab00.place(x=10, y=10)
but01.place(x=10, y=35)
but02.place(x=10, y=60)
but03.place(x=10, y=85)
but01.bind('<Button-1>', filedata_table)
but02.bind('<Button-1>', create_files)
but03.bind('<Button-1>', create_sheets)

root.mainloop()

  
