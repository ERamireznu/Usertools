#v00: 06/07/25

import openpyxl
from openpyxl import Workbook, load_workbook
##import pandas as pd
import datetime
from datetime import timedelta, datetime

def path_parts(pat0):   #reads a path, separates dir and file
    if pat0[-1]=='\\': #a dir ending in \
        return (pat0[:-1], '')
    Lfname = []
    for i in range(-1, -len(pat0)-1,-1):
        if pat0[i]=='\\':
            break
        else:
            Lfname.insert(0, pat0[i])
    fname = ''.join(Lfname)
    fubic = pat0[:len(pat0)+i]
    if '.' not in fname:  #it's a dir
        fname = ''
        fubic = pat0
    return (fubic, fname)

path_def = "C:\\Users\\esteb\\xls_exp.xlsx"
wb = openpyxl.load_workbook(path_def)

#display
salida = False
while not salida:
    print('')
    print('-'*54)
    print("(1)File Data  (2)Create Files  \n(3)Create Sheets  (9)Exit")
    print('option?:')
    opcion = int(input())

    #Sheets' data:        
    if opcion == 1:
        Lfname = []
        print('insert file path (n: default)')
        opt1 = input()
        if opt1.lower() == 'n':
            path = path_def                    
        else:
            path = opt1
        path_exp = path_parts(path)
        fubic_, fname_ = path_exp[0], path_exp[1]
        wb = load_workbook(path, data_only=False)  #considers formulas
        print(f"File : {fname_}")
        print(f"Ubic : {fubic_}")
        print(f"#Shts: {len(wb.sheetnames)}")
        print("show sheets' data? (y/n)")
        ans3 = input()
        if ans3.lower()=='y':
            Res = []
            for hoja in wb.sheetnames:
                ws = wb[hoja]
                used_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
                Res.append((hoja, used_cells))
            print('Sheets ----------')
            print('Name, #cells used:')
            for x in Res:
                print(x[0],', ',x[1])
        else:
            print('-----')

    #Create files:
    elif opcion == 2:
        print('insert dir path (n: default)')
        opt2 = input()
        if opt2.lower() == 'n':
            path = path_def                    
        else:
            path = opt2        
        path_exp = path_parts(path)
        fubic_ = path_exp[0]
        print("enter files' names")
        sh_list = input().split()
        # make files from a list:
        for nom in sh_list:    
            wb = openpyxl.Workbook()
            fpath = fubic_ + '\\' + str(nom) +".xlsx"
            wb.save(fpath)            

    #Create sheets:
    elif opcion == 3:
        print('insert dir path (n: default)')
        opt3 = input()
        if opt3.lower() == 'n':
            path = path_def                    
        else:
            path = opt3

        path_exp = path_parts(path)
        fubic_, fname_ = path_exp[0], path_exp[1]
        
        print("enter sheets' names (new file)")
        sh_list = input().split()
        # make sheets from a list:
        wb = openpyxl.Workbook()
        sheet = wb.active        
        for i in range(len(sh_list)):
            wb.create_sheet(index = i, title = sh_list[i])
        import datetime
        now = datetime.datetime.now()
        dt = now.strftime('%d%m%y_%H%M')
        nom = 'new_' + dt +'.xlsx'        
        wb.save(fubic_ + '\\' + nom)


    #exit:
    elif opcion == 9:
        print('bye '+'/\\'*22)
        salida = True        
